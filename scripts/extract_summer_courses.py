from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = Path("/Users/sophienie/Downloads/暑假课程排课系统.xlsx")
DEFAULT_OUTPUT = ROOT / "src" / "summerCourseLessons.js"

TEACHER_ALIASES = {
    "张佳敏": ("phebe", "Phebe"),
    "Phebe": ("phebe", "Phebe"),
    "Sophie": ("sophie", "Sophie"),
}

COURSE_KEYS = {
    "财商x樱桃 徐汇暑期课": "finance-xuhui",
    "财商x樱桃 浦东暑期课": "finance-pudong",
    "WAICY 徐汇集训班": "waicy-xuhui",
}


def main() -> None:
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SOURCE
    output = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUTPUT
    lessons = expand_lessons(source)
    write_js(lessons, output)
    print(f"Wrote {len(lessons)} summer course lessons to {output}")


def expand_lessons(source: Path) -> list[dict]:
    workbook = load_workbook(source, data_only=True)
    data_ranges = build_data_ranges(workbook)
    lessons = []

    for row_number, row in iter_course_rows(workbook):
        course_name = row.get("课程名称")
        start_date = parse_date(row.get("开始时间"))
        end_date = parse_date(row.get("结束时间"))
        if not course_name or start_date is None or end_date is None:
            continue

        end_date = data_ranges.get((course_name, start_date.isoformat()), end_date)
        teacher_id, teacher_name = map_teacher(row.get("授课老师"))
        start_time, end_time = parse_time_range(row.get("上课时间"))
        location = normalize_text(row.get("上课地点"))
        campus = infer_campus(course_name, location)
        status = normalize_text(row.get("课程状态")) or "Excel导入"
        notes = normalize_text(row.get("备注"))
        student_name = normalize_text(row.get("学员信息")) or "班课"

        for lesson_date in expand_course_dates(course_name, notes, start_date, end_date):
            lessons.append(
                {
                    "id": build_lesson_id(course_name, teacher_id, lesson_date, start_time),
                    "teacherId": teacher_id,
                    "teacherName": teacher_name,
                    "studentName": student_name,
                    "course": course_name,
                    "deliveryType": "线下",
                    "campus": campus,
                    "location": location,
                    "date": lesson_date.isoformat(),
                    "startTime": start_time,
                    "endTime": end_time,
                    "status": status,
                    "notes": notes,
                    "source": f"暑假课程排课系统.xlsx/课程表!A{row_number}",
                }
            )

    return sorted(lessons, key=lambda lesson: (lesson["date"], lesson["startTime"], lesson["teacherName"]))


def build_data_ranges(workbook) -> dict[tuple[str, str], date]:
    if "数据表" not in workbook.sheetnames:
        return {}

    worksheet = workbook["数据表"]
    header = [normalize_text(cell.value) for cell in worksheet[1]]
    ranges = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        values = dict(zip(header, row))
        course_name = normalize_text(values.get("文本"))
        start_date = parse_date(values.get("开始日期"))
        end_date = parse_date(values.get("结束日期"))
        if course_name and start_date and end_date:
            ranges[(course_name, start_date.isoformat())] = end_date
    return ranges


def iter_course_rows(workbook):
    worksheet = workbook["课程表"]
    header = [normalize_text(cell.value) for cell in worksheet[1]]
    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        yield row_number, {key: value for key, value in zip(header, row)}


def expand_course_dates(course_name: str, notes: str, start_date: date, end_date: date) -> list[date]:
    if course_name == "WAICY 徐汇集训班":
        weekdays = {1, 3, 4}
    elif "8天" in notes:
        weekdays = {2, 3, 4, 5}
    else:
        weekdays = {cursor for cursor in range(1, 8)}

    dates = []
    cursor = start_date
    while cursor <= end_date:
        if cursor.isoweekday() in weekdays:
            dates.append(cursor)
        cursor += timedelta(days=1)
    return dates


def parse_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = normalize_text(value)
    if not text:
        return None
    text = text.split("T", 1)[0]
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def parse_time_range(value) -> tuple[str, str]:
    text = normalize_text(value).replace("：", ":").replace("－", "-").replace("—", "-")
    match = re.match(r"^(\d{1,2}):(\d{2})\s*-\s*(\d{1,2}):(\d{2})$", text)
    if not match:
        raise ValueError(f"Invalid time range: {value!r}")
    return format_time(match.group(1), match.group(2)), format_time(match.group(3), match.group(4))


def format_time(hour: str, minute: str) -> str:
    return f"{int(hour):02d}:{int(minute):02d}"


def map_teacher(value) -> tuple[str, str]:
    teacher_name = normalize_text(value)
    if teacher_name not in TEACHER_ALIASES:
        raise ValueError(f"Unknown teacher: {teacher_name}")
    return TEACHER_ALIASES[teacher_name]


def infer_campus(course_name: str, location: str) -> str:
    if "徐汇" in course_name or "上海电影厂" in location:
        return "徐汇"
    if "浦东" in course_name or "樱桃图书馆" in location:
        return "八佰伴"
    return ""


def build_lesson_id(course_name: str, teacher_id: str, lesson_date: date, start_time: str) -> str:
    course_key = COURSE_KEYS.get(course_name, slug(course_name))
    return f"summer-{course_key}-{teacher_id}-{lesson_date.isoformat()}-{start_time.replace(':', '')}"


def slug(value: str) -> str:
    slugged = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "-", value).strip("-")
    return slugged[:32] or "course"


def normalize_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def write_js(lessons: list[dict], output: Path) -> None:
    output.write_text(
        "export const summerCourseLessons = "
        + json.dumps(lessons, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
