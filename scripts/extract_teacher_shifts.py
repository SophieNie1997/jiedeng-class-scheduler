from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE_DIR = ROOT / "老师排班"
DEFAULT_OUTPUT = ROOT / "src" / "importedTeacherShifts.js"
SHEET_NAME = "连续日期排班表"
HEADER_ROW = 4
FIRST_TEACHER_COLUMN = 4

DEFAULT_GRADES = ["G2", "Y3", "Y6", "Y8", "Y9", "大班"]
DEFAULT_DELIVERY_TYPES = ["线上", "线下", "上门", "校区", "樱桃"]
PREFERRED_TEACHER_ORDER = [
    "Claire",
    "Phebe",
    "Sophie",
    "Lynn",
    "Tiana",
    "Catherine",
    "Gioia",
    "Karen",
    "Charlotte",
    "Hanna",
    "Reece",
]


def main() -> None:
    roster, shifts, summary = extract_teacher_shifts_from_sources(discover_shift_sources(DEFAULT_SOURCE_DIR))
    write_js(roster, shifts, summary, DEFAULT_OUTPUT)
    print(f"Wrote {len(roster)} teachers and {len(shifts)} shifts to {DEFAULT_OUTPUT}")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


def discover_shift_sources(folder: Path = DEFAULT_SOURCE_DIR) -> list[Path]:
    return sorted(
        [
            path
            for path in folder.glob("*.xlsx")
            if path.is_file() and not path.name.startswith("~$") and not path.name.startswith(".")
        ],
        key=lambda path: path.name,
    )


def extract_teacher_shifts_from_sources(sources: list[Path]) -> tuple[list[dict], dict[str, dict], dict]:
    roster_by_id = {}
    shifts: dict[str, dict] = {}
    summaries = []
    filled_dates = []
    total_date_rows = 0

    for source in sources:
        roster, source_shifts, summary = extract_teacher_shifts(source)
        summaries.append(summary)
        total_date_rows += summary["dateRows"]
        if summary["filledStartDate"]:
            filled_dates.append(summary["filledStartDate"])
        if summary["filledEndDate"]:
            filled_dates.append(summary["filledEndDate"])
        for teacher in roster:
            roster_by_id[teacher["id"]] = teacher
        shifts.update(source_shifts)

    preferred_ids = [make_teacher_id(name) for name in PREFERRED_TEACHER_ORDER]
    roster = [roster_by_id[teacher_id] for teacher_id in preferred_ids if teacher_id in roster_by_id]
    roster.extend(
        teacher
        for teacher_id, teacher in sorted(roster_by_id.items(), key=lambda item: item[1]["name"].lower())
        if teacher_id not in preferred_ids
    )

    summary = {
        "source": "老师排班",
        "sources": [source.name for source in sources],
        "sheet": SHEET_NAME,
        "teacherCount": len(roster),
        "shiftCount": len(shifts),
        "dateRows": total_date_rows,
        "filledStartDate": min(filled_dates) if filled_dates else "",
        "filledEndDate": max(filled_dates) if filled_dates else "",
        "sourceSummaries": summaries,
    }
    return roster, dict(sorted(shifts.items())), summary


def extract_teacher_shifts(source: Path) -> tuple[list[dict], dict[str, dict], dict]:
    workbook = load_workbook(source, data_only=True)
    worksheet = workbook[SHEET_NAME]
    teacher_columns = find_teacher_columns(worksheet)
    roster = [build_teacher(name) for _, name in teacher_columns]
    shifts = {}
    filled_dates = []
    date_rows = 0

    for row in range(HEADER_ROW + 1, worksheet.max_row + 1):
        shift_date = parse_date(worksheet.cell(row, 2).value)
        if shift_date is None:
            continue
        date_rows += 1

        row_has_shift = False
        for column, teacher_name in teacher_columns:
            shift = parse_shift_value(worksheet.cell(row, column).value)
            if shift is None:
                continue

            row_has_shift = True
            shifts[f"{make_teacher_id(teacher_name)}__{shift_date.isoformat()}"] = shift

        if row_has_shift:
            filled_dates.append(shift_date.isoformat())

    summary = {
        "source": source.name,
        "sheet": SHEET_NAME,
        "teacherCount": len(roster),
        "shiftCount": len(shifts),
        "dateRows": date_rows,
        "filledStartDate": min(filled_dates) if filled_dates else "",
        "filledEndDate": max(filled_dates) if filled_dates else "",
    }
    return roster, dict(sorted(shifts.items())), summary


def find_teacher_columns(worksheet) -> list[tuple[int, str]]:
    columns = []
    for column in range(FIRST_TEACHER_COLUMN, worksheet.max_column + 1):
        name = normalize_text(worksheet.cell(HEADER_ROW, column).value)
        if name:
            columns.append((column, name))
    return columns


def build_teacher(name: str) -> dict:
    return {
        "id": make_teacher_id(name),
        "name": name,
        "courses": [],
        "grades": DEFAULT_GRADES,
        "deliveryTypes": DEFAULT_DELIVERY_TYPES,
        "maxWeeklyHours": 24,
        "weeklyAvailability": [],
        "unavailable": [],
    }


def parse_shift_value(value) -> dict | None:
    label = normalize_text(value)
    if not label:
        return None

    if label == "法定":
        return {"type": "holiday", "label": "法定"}

    if label in {"休", "本休"}:
        return {"type": "off", "label": "休"}

    campus = "徐汇" if label.startswith("徐汇") else "八佰伴"
    time_text = label.replace("徐汇", "").replace("浦东", "").replace("八佰伴", "").replace("碧云", "").replace("早", "").replace("晚", "").strip()
    start_time, end_time = parse_time_range(time_text, prefer_evening=label.startswith("晚"))
    return {
        "type": "work",
        "label": label,
        "startTime": start_time,
        "endTime": end_time,
        "campus": campus,
    }


def parse_time_range(value: str, prefer_evening: bool = False) -> tuple[str, str]:
    match = re.match(r"^(\d{1,2})(?::(\d{2}))?-(\d{1,2})(?::(\d{2}))?$", value)
    if not match:
        raise ValueError(f"Unsupported shift time range: {value!r}")

    start_hour = int(match.group(1))
    start_minute = int(match.group(2) or "0")
    end_hour = int(match.group(3))
    end_minute = int(match.group(4) or "0")

    if prefer_evening and start_hour < 12:
        start_hour += 12
    if end_hour <= start_hour and end_hour < 12:
        end_hour += 12
    if end_hour <= start_hour:
        end_hour += 12

    return format_time(start_hour, start_minute), format_time(end_hour, end_minute)


def parse_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def format_time(hour: int, minute: int) -> str:
    return f"{hour:02d}:{minute:02d}"


def make_teacher_id(name: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "-", name.strip().lower()).strip("-")
    return normalized or "teacher"


def normalize_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def write_js(roster: list[dict], shifts: dict[str, dict], summary: dict, output: Path) -> None:
    output.write_text(
        "export const importedShiftRoster = "
        + json.dumps(roster, ensure_ascii=False, indent=2)
        + ";\n\nexport const importedDefaultShiftOverrides = "
        + json.dumps(shifts, ensure_ascii=False, indent=2)
        + ";\n\nexport const importedShiftSummary = "
        + json.dumps(summary, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
