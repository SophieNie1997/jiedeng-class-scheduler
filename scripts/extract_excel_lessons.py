from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "桔灯 陪伴老师 排课统计.xlsx"
OUTPUT = ROOT / "src" / "importedLessons.js"
YEAR = 2026

TEACHER_SHEETS = {
    "Tiana": "tiana",
    "Lynn": "lynn",
    "Catherine": "catherine",
    "Gioia": "gioia",
    "Karen 7-8月": "karen",
    "Charlotte 7-8月": "charlotte",
    "Hanna 7-8月": "hanna",
    "Hanna 7-8月 ": "hanna",
    "Reece 7-8月": "reece",
}

IGNORE_VALUES = {
    "上午",
    "下午",
    "午休",
    "时间",
    "周日",
    "星期日",
    "星期一",
    "星期二",
    "星期三",
    "星期四",
    "星期五",
    "星期六",
}

GRAY_UNAVAILABLE_COURSE = "灰色不可排"


def main() -> None:
    workbook = load_workbook(SOURCE, data_only=True)
    lessons = []

    for sheet_name, teacher_id in TEACHER_SHEETS.items():
        if sheet_name not in workbook.sheetnames:
            continue

        worksheet = workbook[sheet_name]
        teacher_name = clean_teacher_name(sheet_name)
        blocks = find_schedule_blocks(worksheet)
        merged_lookup = build_merged_lookup(worksheet)
        header_occurrences: dict[tuple[str, int], int] = defaultdict(int)

        for block in blocks:
            header_text = block["header"]
            parsed_range = parse_date_range(header_text)
            if parsed_range is None:
                continue

            occurrence_key = (header_text, block["start_col"])
            occurrence_index = header_occurrences[occurrence_key]
            header_occurrences[occurrence_key] += 1

            for item in iter_block_items(worksheet, block, merged_lookup):
                lesson_date = pick_date_for_weekday(
                    parsed_range[0],
                    parsed_range[1],
                    item["weekday"],
                    occurrence_index,
                )
                if lesson_date is None:
                    continue

                status = "不可用" if item["unavailable"] else infer_status(item["course"])
                lesson_id = (
                    f"excel-{teacher_id}-{lesson_date.isoformat()}-"
                    f"{item['start_time'].replace(':', '')}-{slug(item['course'])}"
                )
                lessons.append(
                    {
                        "id": lesson_id,
                        "teacherId": teacher_id,
                        "teacherName": teacher_name,
                        "studentName": infer_student_name(item["course"], status),
                        "course": item["course"],
                        "deliveryType": infer_delivery_type(item["course"]),
                        "date": lesson_date.isoformat(),
                        "startTime": item["start_time"],
                        "endTime": item["end_time"],
                        "status": status,
                        "source": f"{sheet_name}!{item['coordinate']}",
                    }
                )

    lessons = dedupe_lessons(lessons)
    OUTPUT.write_text(
        "export const importedLessons = "
        + json.dumps(lessons, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )

    by_month: dict[str, int] = defaultdict(int)
    for lesson in lessons:
        by_month[lesson["date"][:7]] += 1

    print(f"Wrote {len(lessons)} imported lessons to {OUTPUT}")
    print(json.dumps(dict(sorted(by_month.items())), ensure_ascii=False, indent=2))


def clean_teacher_name(sheet_name: str) -> str:
    return sheet_name.replace("7-8月", "").strip()


def find_schedule_blocks(worksheet):
    blocks = []
    for row in range(1, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row, col).value
            if not isinstance(value, str):
                continue
            value = normalize_text(value)
            if parse_date_range(value) is None:
                continue
            if not looks_like_weekday_header_row(worksheet, row + 1, col):
                continue

            blocks.append(
                {
                    "header": value,
                    "header_row": row,
                    "start_col": col,
                    "end_col": min(col + 8, worksheet.max_column),
                    "time_col": col + 1,
                }
            )

    blocks.sort(key=lambda block: (block["header_row"], block["start_col"]))

    for index, block in enumerate(blocks):
        next_rows = [
            other["header_row"]
            for other in blocks
            if other["start_col"] == block["start_col"] and other["header_row"] > block["header_row"]
        ]
        block["end_row"] = min(next_rows) - 1 if next_rows else min(block["header_row"] + 27, worksheet.max_row)

    return blocks


def looks_like_weekday_header_row(worksheet, row: int, start_col: int) -> bool:
    values = [normalize_text(worksheet.cell(row, start_col + offset).value) for offset in range(0, 9)]
    joined = " ".join(values)
    return "时间" in joined and "星期一" in joined and ("周日" in joined or "星期日" in joined)


def build_merged_lookup(worksheet):
    lookup = {}
    top_left = {}
    for merged in worksheet.merged_cells.ranges:
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                lookup[(row, col)] = merged
        top_left[(merged.min_row, merged.min_col)] = merged
    return {"all": lookup, "top_left": top_left}


def iter_block_items(worksheet, block, merged_lookup):
    items = []
    seen_ranges = set()

    for row in range(block["header_row"] + 2, block["end_row"] + 1):
        start_end = parse_time_range(worksheet.cell(row, block["time_col"]).value)
        if start_end is None:
            continue

        for col in range(block["start_col"] + 2, block["start_col"] + 9):
            merged = merged_lookup["all"].get((row, col))
            if merged is not None:
                if (merged.min_row, merged.min_col) != (row, col):
                    continue
                range_key = str(merged)
                if range_key in seen_ranges:
                    continue
                seen_ranges.add(range_key)
                cell = worksheet.cell(merged.min_row, merged.min_col)
                value = cell.value
                unavailable = is_gray_fill(cell)
                end_time = parse_time_range(worksheet.cell(merged.max_row, block["time_col"]).value)
                coordinate = str(merged)
                item_start_end = start_end if end_time is None else (start_end[0], end_time[1])
            else:
                cell = worksheet.cell(row, col)
                value = cell.value
                unavailable = is_gray_fill(cell)
                coordinate = f"{get_column_letter(col)}{row}"
                item_start_end = start_end

            course = normalize_text(value)
            if not course and unavailable:
                course = GRAY_UNAVAILABLE_COURSE
            if should_skip_value(course) and not unavailable:
                continue

            weekday = col - (block["start_col"] + 1)
            items.append(
                {
                    "coordinate": coordinate,
                    "weekday": weekday,
                    "course": course,
                    "start_time": item_start_end[0],
                    "end_time": item_start_end[1],
                    "unavailable": unavailable,
                }
            )

    return items


def parse_date_range(text: str):
    text = normalize_text(text)
    patterns = [
        r"(?P<sm>\d{1,2})月(?P<sd>\d{1,2})\s*-\s*(?:(?P<em>\d{1,2})月)?(?P<ed>\d{1,2})",
        r"(?P<sm>\d{1,2})\.(?P<sd>\d{1,2})\s*-\s*(?:(?P<em>\d{1,2})\.)?(?P<ed>\d{1,2})",
    ]

    for pattern in patterns:
        match = re.search(pattern, text)
        if not match:
            continue

        start_month = int(match.group("sm"))
        start_day = int(match.group("sd"))
        end_month = int(match.group("em") or start_month)
        end_day = int(match.group("ed"))
        start = date(YEAR, start_month, start_day)
        end_year = YEAR + 1 if end_month < start_month else YEAR
        end = date(end_year, end_month, end_day)
        return start, end

    return None


def pick_date_for_weekday(start: date, end: date, weekday: int, occurrence_index: int):
    candidates = []
    cursor = start
    while cursor <= end:
        if cursor.isoweekday() == weekday:
            candidates.append(cursor)
        cursor += timedelta(days=1)

    if not candidates:
        return None

    return candidates[min(occurrence_index, len(candidates) - 1)]


def parse_time_range(value):
    text = normalize_text(value)
    match = re.match(r"^(\d{1,2}):(\d{2})\s*-\s*(\d{1,2}):(\d{2})$", text)
    if not match:
        return None
    return format_time(match.group(1), match.group(2)), format_time(match.group(3), match.group(4))


def format_time(hour: str, minute: str) -> str:
    return f"{int(hour):02d}:{int(minute):02d}"


def normalize_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def should_skip_value(value: str) -> bool:
    if not value or value in IGNORE_VALUES:
        return True
    if value.startswith("星期"):
        return True
    if value in {"周六", "周日", "/"}:
        return True
    return False


def is_gray_fill(cell) -> bool:
    fill = cell.fill
    if fill.fill_type != "solid":
        return False

    rgb = getattr(fill.fgColor, "rgb", None)
    if not isinstance(rgb, str) or len(rgb) not in {6, 8}:
        return False

    hex_value = rgb[-6:]
    try:
        red = int(hex_value[0:2], 16)
        green = int(hex_value[2:4], 16)
        blue = int(hex_value[4:6], 16)
    except ValueError:
        return False

    brightness = (red + green + blue) / 3
    spread = max(red, green, blue) - min(red, green, blue)
    return spread <= 14 and 65 <= brightness <= 245


def infer_student_name(course: str, status: str = "Excel导入") -> str:
    if status == "不可用" and course == GRAY_UNAVAILABLE_COURSE:
        return "不可用"

    cleaned = re.sub(r"(?i)\b1v1\b|\b1V1\b|\b1v2\b|\b1V2\b", "", course)
    cleaned = cleaned.replace("线上课", "").replace("线上", "").replace("复习课", "").strip(" /-")
    if "全天营" in cleaned or "班" in cleaned:
        return "班课"
    if len(cleaned) > 18:
        return cleaned[:18]
    return cleaned or "未命名学员"


def infer_delivery_type(course: str) -> str:
    if "线上" in course:
        return "线上"
    if "上门" in course:
        return "上门"
    return "线下"


def infer_status(course: str) -> str:
    if any(term in course for term in ["休息", "请假", "教研", "假期"]):
        return "不可用"
    return "Excel导入"


def slug(value: str) -> str:
    slugged = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "-", value).strip("-")
    return slugged[:24] or "lesson"


def dedupe_lessons(lessons):
    lessons = merge_adjacent_unavailable_lessons(lessons)
    by_key = {}
    for lesson in lessons:
        key = (
            lesson["teacherId"],
            lesson["date"],
            lesson["startTime"],
            lesson["endTime"],
            lesson["course"],
        )
        by_key[key] = lesson
    return sorted(by_key.values(), key=lambda item: (item["date"], item["startTime"], item["teacherName"]))


def merge_adjacent_unavailable_lessons(lessons):
    sorted_lessons = sorted(
        lessons,
        key=lambda item: (
            item["teacherId"],
            item["date"],
            item["course"],
            item["status"],
            item["startTime"],
            item["endTime"],
        ),
    )
    merged = []

    for lesson in sorted_lessons:
        previous = merged[-1] if merged else None
        if (
            previous
            and previous["status"] == "不可用"
            and lesson["status"] == "不可用"
            and previous["teacherId"] == lesson["teacherId"]
            and previous["date"] == lesson["date"]
            and previous["course"] == lesson["course"]
            and previous["endTime"] == lesson["startTime"]
        ):
            previous["endTime"] = lesson["endTime"]
            previous["source"] = f"{previous['source']};{lesson['source']}"
            continue

        merged.append(dict(lesson))

    return merged


if __name__ == "__main__":
    main()
