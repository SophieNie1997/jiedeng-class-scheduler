from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "学员信息"
OUTPUT = ROOT / "src" / "importedStudents.js"


def main() -> None:
    sources = discover_student_sources(SOURCE_DIR)
    students = extract_students_from_sources(sources)
    write_js(students, OUTPUT)
    print(f"Wrote {len(students)} imported students to {OUTPUT}")
    print(json.dumps({"sources": [source.name for source in sources]}, ensure_ascii=False, indent=2))


def discover_student_sources(folder: Path = SOURCE_DIR) -> list[Path]:
    return sorted(
        [
            path
            for path in folder.glob("*.xlsx")
            if path.is_file() and not path.name.startswith("~$") and not path.name.startswith(".")
        ],
        key=lambda path: path.name,
    )


def extract_students_from_sources(sources: list[Path]) -> list[dict]:
    students = []
    for source in sources:
        students.extend(extract_students_from_workbook(source, public=False))
    return merge_students(students)


def extract_students_from_workbook(source: Path, public: bool = True) -> list[dict]:
    workbook = load_workbook(source, data_only=True)
    students: list[dict] = []

    for worksheet in workbook.worksheets:
        headers = read_headers(worksheet)
        if not headers:
            continue

        for row in range(2, worksheet.max_row + 1):
            values = {header: normalize_text(worksheet.cell(row, column).value) for header, column in headers.items()}
            student = build_student_from_row(values, source.name, worksheet.title, row)
            if student is not None:
                students.append(student)

    return merge_students(students) if public else students


def merge_students(students: list[dict]) -> list[dict]:
    merged: dict[str, dict] = {}
    phone_index: dict[str, str] = {}

    for student in students:
        name_key = normalize_key(student["name"])
        phone_key = normalize_key(student.get("_phone", ""))
        key = phone_index.get(phone_key) if phone_key else ""
        if not key:
            key = name_key
        if not key:
            continue
        merged[key] = merge_student(merged.get(key), student)
        if phone_key:
            phone_index[phone_key] = key

    return sorted([public_student(student) for student in merged.values()], key=lambda item: item["name"].lower())


def read_headers(worksheet) -> dict[str, int]:
    headers = {}
    for column in range(1, worksheet.max_column + 1):
        value = normalize_text(worksheet.cell(1, column).value)
        if value:
            headers[value] = column
    return headers


def build_student_from_row(values: dict[str, str], source_name: str, sheet_name: str, row: int) -> dict | None:
    name = clean_student_name(pick_first(values, ["姓名", "孩子姓名"]))
    if not name:
        return None

    grade = pick_first(values, ["年级"])
    school = pick_first(values, ["学校"])
    school_grade = pick_first(values, ["孩子学校和年级"])
    parsed_school, parsed_grade = split_school_grade(school_grade)

    return compact_student(
        {
            "id": f"student-{slug(name)}",
            "name": name,
            "gender": pick_first(values, ["性别"]),
            "grade": grade or parsed_grade,
            "school": school or parsed_school,
            "businessType": pick_first(values, ["业务类型"]),
            "frequency": pick_first(values, ["上课频率"]),
            "needs": pick_first(values, ["当前主要需求陪伴板块"]),
            "_phone": pick_first(values, ["电话", "电话号码 2"]),
            "source": f"{source_name}:{sheet_name}!{row}",
        }
    )


def split_school_grade(value: str) -> tuple[str, str]:
    text = normalize_text(value)
    if not text:
        return "", ""
    match = re.search(r"\b(?:G|Y)\d+\b|大班|中班|小班", text, flags=re.IGNORECASE)
    if not match:
        return text, ""
    grade = match.group(0).upper()
    school = normalize_text(f"{text[:match.start()]} {text[match.end():]}")
    return school, grade


def merge_student(current: dict | None, incoming: dict) -> dict:
    if current is None:
        return dict(incoming)

    merged = dict(current)
    for key, value in incoming.items():
        if key == "source":
            sources = set(filter(None, [merged.get("source"), value]))
            merged[key] = ";".join(sorted(sources))
        elif key == "name" and value:
            preferred_name = choose_preferred_name(merged.get(key, ""), value)
            merged[key] = preferred_name
            merged["id"] = f"student-{slug(preferred_name)}"
        elif value and not merged.get(key):
            merged[key] = value
    return compact_student(merged)


def compact_student(student: dict) -> dict:
    ordered_keys = ["id", "name", "gender", "grade", "school", "businessType", "frequency", "needs", "_phone", "source"]
    return {key: normalize_text(student.get(key)) for key in ordered_keys if normalize_text(student.get(key))}


def public_student(student: dict) -> dict:
    public = dict(student)
    public.pop("_phone", None)
    return public


def pick_first(values: dict[str, str], keys: list[str]) -> str:
    for key in keys:
        value = values.get(key, "")
        if value:
            return value
    return ""


def normalize_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def clean_student_name(value: str) -> str:
    return normalize_text(value).strip(" /")


def choose_preferred_name(current: str, incoming: str) -> str:
    current = clean_student_name(current)
    incoming = clean_student_name(incoming)
    if not current:
        return incoming
    if not incoming:
        return current
    if has_cjk(current) and not has_cjk(incoming):
        return current
    if has_cjk(incoming) and not has_cjk(current):
        return incoming
    return incoming if len(incoming) > len(current) else current


def has_cjk(value: str) -> bool:
    return any("\u4e00" <= char <= "\u9fff" for char in value)


def normalize_key(value: str) -> str:
    return normalize_text(value).lower()


def slug(value: str) -> str:
    slugged = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "-", value).strip("-")
    return slugged[:40] or "student"


def write_js(students: list[dict], output: Path) -> None:
    output.write_text(
        "export const importedStudents = "
        + json.dumps(students, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
