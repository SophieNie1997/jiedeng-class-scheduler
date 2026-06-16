import sys
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from extract_excel_lessons import (  # noqa: E402
    build_merged_lookup,
    dedupe_lessons,
    find_schedule_blocks,
    infer_status,
    iter_block_items,
)


class ExtractExcelLessonsTest(unittest.TestCase):
    def test_gray_blank_cells_are_imported_as_unavailable(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Demo"
        worksheet["A1"] = "7.1-7.7"
        worksheet["A2"] = "时间"
        worksheet["C2"] = "星期一"
        worksheet["D2"] = "星期二"
        worksheet["E2"] = "星期三"
        worksheet["F2"] = "星期四"
        worksheet["G2"] = "星期五"
        worksheet["H2"] = "星期六"
        worksheet["I2"] = "周日"
        worksheet["B3"] = "14:00-14:30"
        worksheet["B4"] = "14:30-15:00"
        gray_fill = PatternFill("solid", fgColor="FFD9D9D9")
        worksheet["C3"].fill = gray_fill
        worksheet["C4"].fill = gray_fill

        blocks = find_schedule_blocks(worksheet)
        items = iter_block_items(worksheet, blocks[0], build_merged_lookup(worksheet))

        self.assertEqual(len(items), 2)
        self.assertTrue(all(item["unavailable"] for item in items))
        self.assertEqual({item["course"] for item in items}, {"灰色不可排"})

    def test_adjacent_gray_unavailable_lessons_are_merged(self):
        lessons = [
            {
                "id": "a",
                "teacherId": "lynn",
                "teacherName": "Lynn",
                "studentName": "不可用",
                "course": "灰色不可排",
                "deliveryType": "线下",
                "date": "2026-07-01",
                "startTime": "14:00",
                "endTime": "14:30",
                "status": "不可用",
                "source": "Demo!C3",
            },
            {
                "id": "b",
                "teacherId": "lynn",
                "teacherName": "Lynn",
                "studentName": "不可用",
                "course": "灰色不可排",
                "deliveryType": "线下",
                "date": "2026-07-01",
                "startTime": "14:30",
                "endTime": "15:00",
                "status": "不可用",
                "source": "Demo!C4",
            },
        ]

        merged = dedupe_lessons(lessons)

        self.assertEqual(len(merged), 1)
        self.assertEqual(merged[0]["startTime"], "14:00")
        self.assertEqual(merged[0]["endTime"], "15:00")
        self.assertEqual(merged[0]["status"], "不可用")

    def test_holiday_labels_are_imported_as_unavailable(self):
        self.assertEqual(infer_status("端午假期"), "不可用")
        self.assertEqual(infer_status("法定假期"), "不可用")


if __name__ == "__main__":
    unittest.main()
